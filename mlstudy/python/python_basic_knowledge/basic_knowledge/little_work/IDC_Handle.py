import openpyxl
import re
import xlrd

file_path = r'D:\System_files\Desktop\reguar_expression\IDC_handle.xlsx'

#正则表达式,用以匹配所有一级域名(不包含个人,国外域名)
pattern =r'\w+\.(com$|edu$|gov.cn$|mil$|net$|org$|biz$|info$|name$|museum$|cn$|cc$|tv$|fm$|im$|com.cn$)'

# def vlookup(checked, result,io):
#
#     '将以元组形式的参数拆分'
#     checked_table = checked[0]
#     checked_col = checked[1]
#
#     result_table =result[0]
#     result_col = result[1]
#
#     input_col   = io[0]
#     output_col = io[1]
#
#
#
#     for table_row in range(2, checked_table.max_row+1):
#         check_tab = checked_table.cell(row=table_row,column = checked_col).value
#         for i in range(2,result_table.max_row+1):
#            if check_tab == result_table.cell(row = i ,column = result_col).value:
#                checked_table.cell(row = table_row ,column = output_col).value = result_table.cell(row = i, column = input_col).value
#            else:
#                continue

def preprocessing_excel_file(file_path):

    wb = openpyxl.load_workbook(file_path)
    result_table = wb.worksheets[0]
    excel = xlrd.open_workbook(file_path)
    res_tab = excel.sheet_by_name('result')

    print(res_tab.nrows)

    # 在域名后插入一级域名(临时)列
    result_table.insert_cols(8, 1)
    result_table.cell(1, 8).value = 'first_domain_name(temp)'


    # 选中一级域名列,构造一级域名
    for table_row in range(2,res_tab.nrows+1):
        # 构造一级域名
        original_domain_value = result_table.cell(row=table_row, column=2).value
        reg = re.search(pattern, original_domain_value)
        temp_domain = result_table.cell(row=table_row, column=8)
        if reg:
            temp_domain.value = reg.group(0)

    # 对域名进行vlookup
    # checked_domain = (result_table, 3)
    # result_domain = (domain_table, 1)
    # io_domain = (2, 7)
    # vlookup(checked_domain,result_domain,io_domain)
    #
    #对IP进行vlookup
    # checked_ip = (result_table, 1)
    # result_ip = (ip_check_table, 2)
    # io_ip = (3, 6)
    # vlookup(checked_ip, result_ip,io_ip)

    # 删除first_domain_name(temp)列
    # result_table.delete_cols(3)

    wb.save(file_path)

if __name__ == '__main__':
    preprocessing_excel_file(file_path)
    print('Done')