import openpyxl
import time
from datetime import datetime

#包含起始IP和结束IP的源文件
file_path =r'D:\System_files\Desktop\reguar_expression\ip_check.xlsx'

#将IP地址根据起始IP及结束IP进行填充
def ip_fill(start_ip,end_ip):
    for i in range(len(start_ip)):
        ips_first = start_ip[i].split('.')
        ips_last = end_ip[i].split('.')
        if (ips_first[0] == ips_last[0]
            and ips_first[1] == ips_last[1]
            and ips_first[2] == ips_last[2]):
            for x in range(int(ips_first[3]),int(ips_last[3])+1):
                iplist = (str(ips_first[0])+'.'+str(ips_first[1])+'.'
                          +str(ips_first[2])+'.'+str(x))
                result.append ((start_ip[i],iplist))

        else:
            for j in range(int(ips_first[3]),256):
                iplist  = (str(ips_first[0])+'.'+str(ips_first[1])+'.'
                          +str(ips_first[2])+'.'+str(j))
                result.append ((start_ip[i],iplist))

            for k in range(int(ips_first[2])+1,int(ips_last[2])):
                for j in range(0,255):
                    iplist = (str(ips_first[0]) + '.' + str(ips_first[1]) + '.'
                              + str(i) + '.' + str(k))
                    result.append ((start_ip[i],iplist))

            for j in range(0,int(ips_last[3])):
                iplist = (str(ips_first[0]) + '.' + str(ips_first[1]) + '.'
                          + str(ips_last[2]) + '.' + str(j))
                result.append ((start_ip[i],iplist))

if __name__ == '__main__':
    start_time = time.strftime(('%H:%M:%S'))
    time_start = datetime.strptime(start_time, '%H:%M:%S')

    start_ip = []
    end_ip = []
    result = []
    wb = openpyxl.load_workbook(file_path)

    ip_source_table = wb.worksheets[0]
    ip_result_table = wb.worksheets[1]

    for i in range(2, ip_source_table.max_row):
        start_ip_row = ip_source_table.cell(row=i, column=3).value
        end_ip_row = ip_source_table.cell(row=i, column=4).value
        start_ip.append(start_ip_row)
        end_ip.append(end_ip_row)

    ip_fill(start_ip,end_ip)

    for i in range(0,len(result)):
        ip_start = ip_result_table.cell(row=i+1, column=1)
        ip_start.value = result[i][0]
        ip_detail = ip_result_table.cell(row=i+1, column=2)
        ip_detail.value= result[i][1]

        for m in range(0,ip_source_table.max_row):
            if ip_detail.value == ip_source_table.cell(row=m+1, column=3).value:
                ip_result_table.cell(row=i+1, column=3).value = ip_source_table.cell(row=m+1, column=1).value

    ip_result_table.insert_rows(0)
    ip_result_table.cell(1, 1).value = 'ip_start'
    ip_result_table.cell(1, 2).value = 'ip_detail'
    ip_result_table.cell(1, 3).value = 'ec_name'

    end_time = time.strftime(('%H:%M:%S'))
    time_end = datetime.strptime(end_time, '%H:%M:%S')
    print(start_time, end_time, time_end - time_start)

    wb.save(file_path)
