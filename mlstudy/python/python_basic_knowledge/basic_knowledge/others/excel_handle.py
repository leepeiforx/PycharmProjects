

from xlwt import *
import xlrd
#xlwt,xlrd处理03或更低版本文件

book = Workbook()                   #创建新的EXCEL工作簿文件
sheet1 = book.add_sheet('First')    #添加新的worksheet
a1 = Alignment()
a1.horz = Alignment.HORZ_CENTER     #对齐方式
a1.vert = Alignment.VERT_CENTER

borders = Borders()
borders.bottom = Borders.THICK      #边框样式

style = XFStyle()
style.alignment = a1
style.borders = borders

row = sheet1.row(0)                 #获取第0行
row.write(0, 'test', style=style)     #写入单元格
row = sheet1.row(1)
for i in range(5):
    row.write(i, i, style=style)           #写入数字
row.write(5, '=SUM(A2:E2)', style=style)    #写入公式
book.save(r'C:\Users\Administrator\Desktop\da\Excel_handel.xls')

book = xlrd.open_workbook(r'C:\Users\Administrator\Desktop\da\Excel_handel.xls')
sheet1 = book.sheet_by_name('First')
row = sheet1.row(0)
print(row[0].value )
print(sheet1.row(1)[2].value )


import openpyxl
from openpyxl import Workbook
import random

fn = r'C:\Users\Administrator\Desktop\da\excelup.xlsx'          #文件名
wb = Workbook()                                                 #创建工作簿
ws = wb.create_sheet(title='hello,world')                       #创建工作表
ws['A1'] = '第一个单元格'                                         #单元格赋值
ws['B1'] = 3.1415926
wb.save(fn)                                                     #保存Excel文件

wb = openpyxl.load_workbook(fn)                                 #w打开已有的Excel文件
ws = wb.worksheets[1]                                           #打开指定索引的工作表
print(ws['A1'].value)                                           #读取并输出指定单元格的值
ws.append([1,2,3,4,5])                                          #添加一行数据
ws.merge_cells('F2:F3')                                         #合并单元格
ws['F2']='=SUM(A2:E2)'                                          #写入公式
for r in range(10,15):
    for c in range(3,8):
        _ = ws.cell(row=r, column=c, value=r*c)                  #写入单元格数据

wb.save(fn)

def generateRandomInfo(filename):

    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(['姓名', '课程', '成绩'])

    #中文名称中的字
    first = '赵钱孙李'
    middle = '一二三四'
    last = '周吴郑'

    subjects = ('语文', '数学', '英语')

    for i in range(200):
        line = []
        r = random.randint(1, 100)
        name = random.choice(first)
        #按照一定概率生产只有两个字的中文名称
        if r > 50:
            name = name + random.choice(middle)
        name = name + random.choice(last)

        #依次生产姓名,课程名称和成绩
        line.append(name)
        line.append(random.choice(subjects))
        line.append(random.randint(0, 100))

        worksheet.append(line)

    #保存数据,生产xlsx格式文件
    workbook.save(filename)

def getResult(oldfile,newfile):

    #用于存放结果的字典
    result = dict()

    #打开原始数据
    workbook = openpyxl.load_workbook(oldfile)
    worksheet = workbook.worksheets[0]

    #遍历原始数据
    for row in worksheet.rows:
        if row[0].value == '姓名':
            continue
        #姓名,课程名称,本次成绩
        name, subject, grade = row[0].value, row[1].value, row[2].value

        #获取当前姓名对应的课程名称和成绩信息
        #如果result字典中不包含,则返回空字典
        t = result.get(name, {})

        #获取当前学生成绩的课程,若不存在,则返回0
        f = t.get(subject, 0)

        #只保留学生该课程的最高成绩
        if grade > f:
            t[subject] = grade
            result[name] = t

    workbook1 = Workbook()
    worksheet1 = workbook1.worksheets[0]
    worksheet1.append(['姓名', '课程', '成绩'])

    #将result字典的结果写入Excel文件
    for name, t in result.items():
        print(name, t)
        for subject, grade in t.items():
            worksheet1.append([name, subject, grade])

    workbook1.save(newfile)

if __name__ == '__main__':
    oldfile = r'C:\Users\Administrator\Desktop\da\test.xlsx'
    newfile = r'C:\Users\Administrator\Desktop\da\result.xlsx'
    generateRandomInfo(oldfile)
    getResult(oldfile, newfile)

