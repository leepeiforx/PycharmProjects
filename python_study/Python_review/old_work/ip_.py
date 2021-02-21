import more_itertools as mit
import openpyxl
import pandas as pd

iterable = []
result = []
#数据源文件路径
file_path = r'D:\System_files\Desktop\ip.xlsx'

def open_file (file_path):

   wb = openpyxl.load_workbook(file_path)
   ip_source_table = wb.worksheets[0]
   return ip_source_table

#整理连续数并组合成列表
def find_ranges(iterable):

    _list = [list(group) for group in mit.consecutive_groups(iterable)]
    return _list

#返回结果
def get_ranges_detail(_list):

    for _li in _list:
        number_segment = str(min(_li))[0:8]
        detail_number_segment = str(min(_li))[-5:]+'~'+str(max(_li))[-5:]
        consecutive_numbers = len(_li)

        result.append((number_segment, detail_number_segment, consecutive_numbers))

def get_ip_row(table):

    ip_list = []
    for i in range(1, table.max_row):
        ip_list.append(table.cell(row=i, column=1).value)
    return ip_list


if __name__ == '__main__':

    ip_table = open_file(file_path)
    ip_list = get_ip_row(ip_table)
    _list = find_ranges(ip_list)

    get_ranges_detail(_list)

    name = ['号段','细分号段','连续号码数量']
    test = pd.DataFrame(columns=name,data=result)
    test.to_excel(file_path,encoding='gbk')
