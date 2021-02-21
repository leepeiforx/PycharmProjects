import openpyxl
import re

file_path = r'D:\System_files\Desktop\reguar_expression\regular_expression.xlsx'
result = []

#正则表达式,用以匹配所有一级域名(不包含个人,国外域名)
pattern = ('\w+\.(com$|edu$|gov.cn$|mil$|net$|org$|biz$|info$|name$|museum$|cn$|cc$|tv$|fm$|im$|com.cn$)')

def get_regular_expression(str):
    if str is not None:
        match = re.search(pattern, str)
        if match:
            return match.group(0)

if __name__ == '__main__':
    wb = openpyxl.load_workbook(file_path)
    domain_name_table = wb.worksheets[0]

    for i in range(2,domain_name_table.max_row+1):
        domain_name_row = domain_name_table.cell(i, column=1).value
        # print(domain_name_row)
        domain_name_table.cell(i, column=3).value = get_regular_expression(domain_name_row)
    print('done')
    wb.save(file_path)




